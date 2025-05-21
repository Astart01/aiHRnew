import streamlit as st
import pandas as pd
import numpy as np
import os
import re
import string
import joblib
import fitz  # PyMuPDF
import base64
import pymorphy3
import nltk
from nltk.corpus import stopwords
from nltk.stem.snowball import SnowballStemmer
from catboost import CatBoostClassifier
import hashlib
import json
import io
import subprocess
import sys
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pdf2image

# Установка заголовка страницы - это ДОЛЖНА быть первая команда Streamlit
st.set_page_config(
    page_title="Классификатор резюме продавцов",
    page_icon="📄",
    layout="wide"
)

# Проверка параметра очистки в URL
query_params = st.query_params
if "clear" in query_params:
    # Очистка всех сессий на системном уровне
    if hasattr(st, "session_state"):
        for key in list(st.session_state.keys()):
            if key not in ['authenticated', 'user_role', 'user_name']:
                del st.session_state[key]
    
    # Инициализация базовых переменных
    st.session_state.processed_files = {}
    st.session_state.results = []
    st.session_state.has_processed_files = False
    st.session_state.selected_rows = set()
    
    # Удаляем параметр очистки из URL
    st.query_params.clear()

# --- Настройки авторизации ---
USERS_FILE = "users.json"

# --- Функции управления пользователями ---
def load_users():
    if os.path.exists(USERS_FILE):
        with open(USERS_FILE, "r", encoding='utf-8') as f:
            return json.load(f)
    else:
        default_users = {
            "admin": {
                "password": hashlib.sha256("admin".encode()).hexdigest(),
                "role": "admin",
                "name": "Администратор"
            },
            "user": {
                "password": hashlib.sha256("password".encode()).hexdigest(),
                "role": "user",
                "name": "Пользователь"
            }
        }
        with open(USERS_FILE, "w", encoding='utf-8') as f:
            json.dump(default_users, f, ensure_ascii=False, indent=4)
        return default_users

def save_users(users):
    with open(USERS_FILE, "w", encoding='utf-8') as f:
        json.dump(users, f, ensure_ascii=False, indent=4)

def authenticate(username, password):
    users = load_users()
    hashed_password = hashlib.sha256(password.encode()).hexdigest()
    if username in users and users[username]["password"] == hashed_password:
        st.session_state.user_role = users[username]["role"]
        st.session_state.user_name = users[username].get("name", username)
        return True
    return False

# --- Страница администратора ---
def admin_panel():
    st.title("Панель администратора")
    
    tab1, tab2 = st.tabs(["Управление пользователями", "Добавить пользователя"])
    
    with tab1:
        st.subheader("Управление пользователями")
        users = load_users()
        
        user_list = []
        for username, user_data in users.items():
            user_list.append({
                "Логин": username,
                "Имя": user_data.get("name", ""),
                "Роль": user_data.get("role", "user")
            })
        
        user_df = pd.DataFrame(user_list)
        st.dataframe(user_df, use_container_width=True)
        
        # Удаление пользователя
        st.subheader("Удалить пользователя")
        user_to_delete = st.selectbox("Выберите пользователя для удаления", 
                                     [u for u in users.keys() if u != "admin"])
        
        if st.button("Удалить пользователя"):
            if user_to_delete == "admin":
                st.error("Невозможно удалить администратора!")
            else:
                users.pop(user_to_delete, None)
                save_users(users)
                st.success(f"Пользователь {user_to_delete} успешно удален")
                st.rerun()
    
    with tab2:
        st.subheader("Добавить нового пользователя")
        
        with st.form("add_user_form"):
            new_username = st.text_input("Логин")
            new_password = st.text_input("Пароль", type="password")
            new_name = st.text_input("Имя")
            new_role = st.selectbox("Роль", ["user", "admin"])
            
            submit = st.form_submit_button("Добавить пользователя")
            
            if submit:
                users = load_users()
                if new_username in users:
                    st.error(f"Пользователь с логином {new_username} уже существует!")
                elif not new_username or not new_password:
                    st.error("Логин и пароль обязательны для заполнения!")
                else:
                    users[new_username] = {
                        "password": hashlib.sha256(new_password.encode()).hexdigest(),
                        "role": new_role,
                        "name": new_name
                    }
                    save_users(users)
                    st.success(f"Пользователь {new_username} успешно добавлен")

# --- Инициализация сессионного состояния ---
if 'authenticated' not in st.session_state:
    st.session_state.authenticated = False
if 'user_role' not in st.session_state:
    st.session_state.user_role = None
if 'user_name' not in st.session_state:
    st.session_state.user_name = None
if 'processed_files' not in st.session_state:
    st.session_state.processed_files = {}
if 'results' not in st.session_state:
    st.session_state.results = []
if 'has_processed_files' not in st.session_state:
    st.session_state.has_processed_files = False
if 'selected_rows' not in st.session_state:
    st.session_state.selected_rows = set()

# --- Загрузка NLTK данных ---
nltk.download('stopwords')

# --- Инициализация морфологического анализатора и стеммера ---
morph = pymorphy3.MorphAnalyzer()
stemmer = SnowballStemmer("russian")
default_stopwords = set(stopwords.words("russian"))
custom_stopwords = default_stopwords - {"без", "для", "по", "при", "над"}
custom_stopwords |= {
    "резюме", "обновлено", "контакт", "зарплата", "телефон", "месяц", "лет",
    "января", "февраля", "марта", "апреля", "мая", "июня", "июля", "августа",
    "сентября", "октября", "ноября", "декабря", "должность", "работа",
    "компания", "обязанности", "основной", "задача", "опыт", "место",
    "года", "году", "владение", "информация", "образование", "гражданство"
}

# --- Ручные фичи ---
features = {
    "sales_experience": [
        r"опыт.*прода", r"звонк[а-я]*", r"\bCRM\b", r"SPIN", r"AIDA", r"скрипт",
        r"обработка заявк", r"воронк", r"телемаркет", r"менеджер по продаж"
    ],
    "hard_skills": [
        r"возражен", r"переговор", r"следовать инструкц", r"многозада",
        r"ведение.*переговор", r"обработка входящ", r"1с", r"excel", r"анализ"
    ],
    "soft_skills": [
        r"мотивир", r"самостоят", r"проактив", r"обуча[еия]", r"дружелюб", r"стрессоустойчив",
        r"клиентоориент", r"гибкост", r"адаптир", r"энергичн", r"настойчив", r"коммуникаб"
    ],
    "performance_metrics": [
        r"конверс", r"выручк", r"чек", r"план", r"результат", r"рост.*конверс",
        r"лояльн", r"возврат", r"kpi", r"достиж", r"удвоил", r"выполн", r"закрыт"
    ]
}

# --- Функции для анализа резюме и комментирования (перенесены из comments.py) ---
def detect_red_flag_areas(text):
    text = text.lower()
    
    # Нежелательные области опыта
    red_flag_areas = {
        "фитнес": ["фитнес", "тренер", "спортивный клуб", "фитнес центр", "тренажерный зал"],
        "недвижимость": ["недвижимость", "риэлтор", "агент по недвижимости", "агентство недвижимости", "продажа квартир", "продажа домов"],
        "авто": ["автосалон", "автомобили", "машины", "продажа авто", "автодилер", "продавец авто"],
        "банки": ["банк", "кредитный специалист", "кредитный менеджер", "финансовый консультант", "ипотечный", "кредитные продукты"],
        "салоны красоты": ["салон красоты", "косметика", "парикмахер", "стилист", "визажист", "косметолог"],
        "продавец-консультант": ["продавец-консультант", "консультант по продажам", "продавец в магазине", "консультация покупателей", "работа в торговом зале"]
    }
    
    # Положительные индикаторы (телефонные продажи)
    positive_indicators = [
        "телефонные продажи", "холодные звонки", "телемаркетинг", "продажи по телефону",
        "call-центр", "колл центр", "телефонные переговоры", "обзвон клиентов",
        "холодная база", "лиды", "удаленные продажи", "оператор call-центра"
    ]
    
    # Проверяем совпадения с red flag областями
    found_red_flags = []
    for area, keywords in red_flag_areas.items():
        for keyword in keywords:
            if keyword in text:
                found_red_flags.append(area)
                break
    
    # Проверяем положительные индикаторы
    has_phone_sales = False
    for indicator in positive_indicators:
        if indicator in text:
            has_phone_sales = True
            break
    
    # Если есть red flags, но нет телефонных продаж
    if found_red_flags and not has_phone_sales:
        return True, found_red_flags, has_phone_sales
    
    return False, found_red_flags, has_phone_sales

def get_detailed_comment(text, predicted_class, relevance_prob):
    # Проверяем red flags
    is_red_flag, red_flag_areas, has_phone_sales = detect_red_flag_areas(text)
    
    # Начинаем с пустого комментария
    comment = ""
    
    if is_red_flag:
        red_flag_str = ", ".join(set(red_flag_areas))
        comment += f"RED FLAG: Имеет опыт работы в областях: {red_flag_str}, но отсутствует опыт телефонных продаж."
    elif red_flag_areas and has_phone_sales:
        red_flag_str = ", ".join(set(red_flag_areas))
        comment += f"Имеет опыт работы в областях: {red_flag_str}, но присутствует опыт телефонных продаж, что является положительным фактором."
    elif has_phone_sales:
        comment += "Кандидат имеет опыт телефонных продаж, что соответствует требованиям позиции."
    elif predicted_class == 0:
        comment += "Недостаточное соответствие требованиям."
    
    # Анализ наличия ключевых навыков для продаж
    sales_skills = []
    
    skill_patterns = {
        "CRM": ["crm", "срм", "customer relationship management"],
        "Холодные звонки": ["холодные звонки", "холодный обзвон", "холодная база"],
        "Работа с возражениями": ["возражен", "работа с возражениями", "отработка возражений"],
        "Ведение переговоров": ["переговоры", "ведение переговоров", "навыки переговоров"],
        "SPIN/AIDA": ["spin", "aida", "техника продаж", "методы продаж"],
        "Выполнение плана": ["план продаж", "выполнение плана", "перевыполнение плана", "плановые показатели"],
        "Аналитика продаж": ["аналитика", "анализ продаж", "продажная воронка", "конверсия"]
    }
    
    for skill, patterns in skill_patterns.items():
        for pattern in patterns:
            if pattern in text.lower():
                sales_skills.append(skill)
                break
    
    if sales_skills:
        skills_str = ", ".join(sales_skills)
        comment += f" Обладает следующими навыками: {skills_str}."
    else:
        comment += " В резюме не указаны ключевые навыки для телефонных продаж."
    
    return comment, is_red_flag

# --- Загрузка модели и вспомогательных объектов ---
@st.cache_resource
def load_model():
    model_path = "catboost_model.cbm"
    if not os.path.exists(model_path):
        st.error(f"Файл модели {model_path} не найден!")
        return None, None, None
    try:
        model = CatBoostClassifier()
        model.load_model(model_path)
        try:
            scaler = joblib.load("scaler.pkl")
        except:
            scaler = joblib.load("scaler.joblib")
        try:
            tfidf = joblib.load("tfidf_vectorizer.pkl")
        except:
            tfidf = joblib.load("tfidf_vectorizer.joblib")
        st.success("Модель и компоненты успешно загружены")
        return model, scaler, tfidf
    except Exception as e:
        st.error(f"Ошибка при загрузке модели: {e}")
        import traceback
        st.code(traceback.format_exc())
        return None, None, None
    
# --- Функции обработки ---
def extract_text_from_pdf(pdf_file):
    try:
        doc = fitz.open(stream=pdf_file.read(), filetype="pdf")
        text = ""
        for page in doc:
            text += page.get_text()
        doc.close()
        return text
    except Exception as e:
        return f"[Ошибка] {e}"

# Добавьте эту функцию перед функцией display_pdf
def convert_pdf_to_images(pdf_file):
    try:
        pdf_file.seek(0)
        # Значительно уменьшаем DPI для получения изображений меньшего размера
        # Чем меньше DPI, тем меньше будет текст
        images = pdf2image.convert_from_bytes(
            pdf_file.read(),
            dpi=72,  # Уменьшаем DPI до 72 (стандартное разрешение экрана)
            fmt='jpeg',
            size=(800, None)  # Ограничиваем ширину изображения 800 пикселями
        )
        return images
    except Exception as e:
        st.error(f"Ошибка при конвертации PDF в изображения: {e}")
        return None

def display_pdf(pdf_file):
    images = convert_pdf_to_images(pdf_file)
    if images:
        # Создаем вкладки для каждой страницы PDF
        if len(images) > 1:
            tabs = st.tabs([f"Стр. {i+1}" for i in range(len(images))])
            for i, tab in enumerate(tabs):
                with tab:
                    # Создаем изображение с явным контролем размера
                    img_resized = images[i].resize((800, int(800 * images[i].height / images[i].width)))
                    st.image(img_resized, use_container_width=True)
        else:
            # Если только одна страница, просто показываем ее с уменьшенным размером
            img_resized = images[0].resize((800, int(800 * images[0].height / images[0].width)))
            st.image(img_resized, use_container_width=True)
    else:
        # Запасной вариант
        try:
            pdf_file.seek(0)
            base64_pdf = base64.b64encode(pdf_file.read()).decode("utf-8")
            pdf_display = f'<iframe src="data:application/pdf;base64,{base64_pdf}" width="100%" height="600px" type="application/pdf"></iframe>'
            st.markdown(pdf_display, unsafe_allow_html=True)
        except Exception as e:
            st.error(f"Не удалось отобразить PDF: {e}")

def extract_resume_info(text):
    info = {
        "phone": "-",
        "position": "-",
        "city": "-",
        "age": "-",
        "gender": "-",
        "salary": "-"
    }
    phone_match = re.search(r'\+7\s*\(?\d{3}\)?[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}', text)
    if phone_match:
        info["phone"] = phone_match.group()
    position_match = re.search(r"[жЖ]елаемая должность и зарплата\s*[:—]?\s*\s*(.*?)(?=\n|$)", text, re.IGNORECASE | re.DOTALL)
    if position_match:
        info["position"] = position_match.group(1).strip()
    city_match = re.search(r'(Москва|Санкт-Петербург|Екатеринбург|Казань|Новосибирск|Самара|Омск|Челябинск)', text)
    if city_match:
        info["city"] = city_match.group(1)
    age_match = re.search(r',\s(\d{2})\s*(год|лет|года),', text)
    if age_match:
        info["age"] = age_match.group(1)
    if 'женщина,' in text.lower()[:500]:
        info["gender"] = "Женщина"
    elif 'мужчина,' in text.lower()[:500]:
        info["gender"] = "Мужчина"
    salary_match = re.search(r'\d{2,3}\s*(000|т.р.)\s*(₽|р|руб)', text)
    if salary_match:
        info["salary"] = re.sub(r'\D', '', salary_match.group(0))
    return info

def preprocess_resume(text):
    cover_idx = text.lower().find("сопроводительное письмо")
    position_idx = text.lower().find("желаемая должность и зарплата")
    if cover_idx != -1:
        text = text[cover_idx:]
    elif position_idx != -1:
        text = text[position_idx:]
    text = re.sub(r'Сопроводительное письмо', '[COVER]', text, flags=re.IGNORECASE)
    text = re.sub(r'Желаемая должность и зарплата', '[POSITION]', text, flags=re.IGNORECASE)
    text = re.sub(r'Специализации', '[SPECIALIZATIONS]', text, flags=re.IGNORECASE)
    text = re.sub(r'Занятость:.*?Опыт работы —', 'Опыт работы —', text, flags=re.DOTALL)
    text = text.split('История общения с кандидатом')[0]
    text = re.sub(r'\S+@\S+', ' ', text)
    text = re.sub(r'\+7\s*\(?\d{3}\)?[\s\-]?\d{3}[\s\-]?\d{2}[\s\-]?\d{2}', ' ', text)
    text = re.sub(r'http\S+|www\.\S+|\S+\.ru|\S+\.com', ' ', text)
    months = r'(январ[ья]|феврал[ья]|марта?|апрел[ья]|ма[йя]|июн[ья]|июл[ья]|август[а]?|сентябр[ья]|октябр[ья]|ноябр[ья]|декабр[ья])'
    text = re.sub(rf'{months}\s+\d{{4}}\s*[—-]\s*{months}\s+\d{{4}}', ' ', text, flags=re.IGNORECASE)
    text = re.sub(rf'{months}\s+\d{{4}}', ' ', text, flags=re.IGNORECASE)
    text = re.sub(r'\b\d{1,2}[./]\d{1,2}[./]\d{2,4}\b', ' ', text)
    text = re.sub(r'\b\d{4}\b', ' ', text)
    text = re.sub(r'\b\d+\b', ' ', text)
    text = re.sub(r'<.*?>', ' ', text)
    text = text.translate(str.maketrans('', '', string.punctuation + '•–—'))
    replacements = {
        r'Навыки': '[SKILLS]',
        r'Обо мне': '[ABOUT]',
        r'Опыт работы': '[EXPERIENCE]',
        r'Образование': '[EDUCATION]',
        r'Знание языков': '[LANGUAGES]',
        r'Дополнительная информация': '[EXTRA]',
    }
    for pattern, tag in replacements.items():
        text = re.sub(pattern, tag, text, flags=re.IGNORECASE)
    text = text.lower()
    words = re.findall(r'\b\w+\b', text)
    processed_words = []
    for word in words:
        if word not in custom_stopwords:
            lemma = morph.parse(word)[0].normal_form
            stem = stemmer.stem(lemma)
            processed_words.append(stem)
    if len(processed_words) > 2:
        processed_words = processed_words[:-2]
    return ' '.join(processed_words)

def extract_features(text, feature_dict):
    text = text.lower()
    return {
        category: sum(int(re.search(pattern, text) is not None) for pattern in patterns)
        for category, patterns in feature_dict.items()
    }

def extract_resume_features(text):
    clean_text = text.replace('\n', ' ').replace('\r', ' ').lower()
    try:
        features = {
            'gender': 1 if 'женщина,' in clean_text[:500] else (-1 if 'мужчина,' in clean_text[:500] else 0),
            'age': int(re.search(r',\s(\d{2})\s*(год|лет|года),', clean_text).group(1)) if re.search(r',\s(\d{2})\s*(год|лет|года),', clean_text) else -1,
            'salary': int(re.sub(r'\D', '', re.search(r'\d{2,3}\s*(000|т.р.)\s*(₽|р|руб)', clean_text)[0])) if re.search(r'\d{2,3}\s*(000|т.р.)\s*(₽|р|руб)', clean_text) else -1,
            'student': int('студент' in clean_text or 'учусь' in clean_text or 'очная' in clean_text),
            'wants_sales_position': int('продаж' in clean_text),
            'text_length': len(clean_text),
            'num_digits': sum(c.isdigit() for c in clean_text),
        }
        return features
    except Exception as e:
        st.warning(f"Ошибка при извлечении признаков из резюме: {e}")
        return {
            'gender': 0,
            'age': -1,
            'salary': -1,
            'student': 0,
            'wants_sales_position': 0,
            'text_length': len(clean_text),
            'num_digits': sum(c.isdigit() for c in clean_text),
        }

# --- Функция отправки в AmoCRM ---
def send_to_amocrm():
    if not st.session_state.results or len(st.session_state.results) == 0:
        st.error("Нет данных для отправки в AmoCRM. Сначала обработайте файлы.")
        return False
    
    # Создаем временный CSV файл с результатами
    result_df = pd.DataFrame(st.session_state.results)
    result_df["Файл"] = result_df["Файл"].str.replace('.pdf', '', regex=False)
    result_df["Вероятность класса 1"] = result_df["Вероятность класса 1"].astype(float).map("{:.2f}".format)
    result_df["Зарплата"] = result_df["Зарплата"].apply(lambda x: f"{int(x):,}".replace(',', ' ') if str(x).isdigit() else x)
    
    display_df = result_df.drop(columns=["raw_proba", "raw_text", "prediction_class"], errors="ignore")
    
    # Сохраняем DataFrame в CSV
    temp_csv_path = "temp_results.csv"
    display_df.to_csv(temp_csv_path, index=False, encoding='utf-8')
    
    try:
        # Вызываем функцию из amo_script.py
        from amo_script import AmoCRMClient
        
        with st.spinner("Отправка данных в AmoCRM..."):
            client = AmoCRMClient(temp_csv_path)
            client.process_csv()
        
        # Удаляем временный CSV файл
        if os.path.exists(temp_csv_path):
            os.remove(temp_csv_path)
        
        return True
    except Exception as e:
        st.error(f"Ошибка при отправке данных в AmoCRM: {e}")
        import traceback
        st.code(traceback.format_exc())
        return False

# --- Страница авторизации ---
def login_page():
    st.title("Авторизация")
    with st.form("login_form"):
        username = st.text_input("Логин")
        password = st.text_input("Пароль", type="password")
        submitted = st.form_submit_button("Войти")
        if submitted:
            if authenticate(username, password):
                st.session_state.authenticated = True
                st.success("Авторизация успешна!")
                st.rerun()
            else:
                st.error("Неверный логин или пароль")

# --- Основное приложение ---
def main_app():
    model, scaler, tfidf = load_model()
    st.title("Классификация резюме менеджеров по продажам")
    st.markdown("### Фокус: поиск кандидатов с опытом телефонных продаж")
    
    # Инициализация состояния для выбранных строк
    if 'selected_rows' not in st.session_state:
        st.session_state.selected_rows = set()
    
    # Обработка кнопки сброса выбранного PDF (если она была нажата)
    if hasattr(st.session_state, 'reset_pdf') and st.session_state.reset_pdf:
        if hasattr(st.session_state, 'selected_pdf'):
            del st.session_state.selected_pdf
        st.session_state.reset_pdf = False
    
    THRESHOLD = 0.19
    uploaded_files = st.file_uploader("Загрузите PDF-файлы", type="pdf", accept_multiple_files=True)
    
    if uploaded_files and st.button("Обработать файлы"):
        results = []
        with st.spinner("Обработка файлов..."):
            for file in uploaded_files:
                file.seek(0)
                raw_text = extract_text_from_pdf(file)
                if "[Ошибка]" in raw_text:
                    results.append({
                        "Файл": file.name,
                        "Вероятность класса 1": 0,
                        "Телефон": "-",
                        "Желаемая должность": "-",
                        "Город": "-",
                        "Возраст": "-",
                        "Пол": "-",
                        "Зарплата": "-",
                        "Комментарий": f"Ошибка обработки файла: {raw_text}"
                    })
                    continue
                st.session_state.processed_files[file.name] = {
                    "file": file,
                    "raw_text": raw_text
                }
                info = extract_resume_info(raw_text)
                processed_text = preprocess_resume(raw_text)
                keyword_features = extract_features(processed_text, features)
                resume_features = extract_resume_features(raw_text)
                manual_df = pd.DataFrame([keyword_features | resume_features])
                tfidf_features = tfidf.transform([processed_text]).toarray()
                combined_features = np.hstack([manual_df.values, tfidf_features])
                scaled_features = scaler.transform(combined_features)
                proba = model.predict_proba(scaled_features)[0]
                raw_proba = proba[1]
                prediction = 1 if raw_proba >= THRESHOLD else 0
                comment, is_red_flag = get_detailed_comment(raw_text, prediction, raw_proba)
                
                results.append({
                    "Файл": file.name,
                    "Вероятность класса 1": raw_proba,
                    "Телефон": info["phone"],
                    "Желаемая должность": info["position"],
                    "Город": info["city"],
                    "Возраст": info["age"],
                    "Пол": info["gender"],
                    "Зарплата": info["salary"],
                    "Комментарий": comment,
                    "raw_proba": raw_proba,
                    "raw_text": raw_text
                })
        st.session_state.results = results
        st.session_state.has_processed_files = True
    
    if st.session_state.has_processed_files and st.session_state.results:
        for r in st.session_state.results:
            r["prediction_class"] = 1 if r["raw_proba"] >= THRESHOLD else 0
            if "raw_text" in r:
                r["Комментарий"], is_red_flag = get_detailed_comment(r["raw_text"], r["prediction_class"], r["raw_proba"])
        
        result_df = pd.DataFrame(st.session_state.results)
        result_df["Файл"] = result_df["Файл"].str.replace('.pdf', '', regex=False)
        result_df["Вероятность класса 1"] = result_df["Вероятность класса 1"].astype(float).map("{:.2f}".format)
        result_df["Зарплата"] = result_df["Зарплата"].apply(lambda x: f"{int(x):,}".replace(',', ' ') if str(x).isdigit() else x)
        
        # Создаем контейнер для результатов
        st.write("### Результаты анализа")
    
        # Добавляем чекбокс "Выделить всех"
        col1, col2, col3 = st.columns([1, 1, 1])
        with col1:
            if st.checkbox("Выделить всех", key="select_all"):
                st.session_state.selected_rows = set(range(len(result_df)))
            else:
                if len(st.session_state.selected_rows) == len(result_df):  # Если все выбраны, снимаем выбор
                    st.session_state.selected_rows = set()
    
        # Добавляем фильтры по цветовым категориям
        with col2:
            if st.checkbox("Зеленые (≥81%)", key="green_filter"):
                green_indices = set(result_df[result_df["Вероятность класса 1"].astype(float) >= 0.81].index)
                st.session_state.selected_rows = st.session_state.selected_rows.union(green_indices)
    
        with col3:
            if st.checkbox("Желтые (19-80%)", key="yellow_filter"):
                yellow_indices = set(result_df[
                    (result_df["Вероятность класса 1"].astype(float) >= 0.19) & 
                    (result_df["Вероятность класса 1"].astype(float) < 0.81)
                ].index)
                st.session_state.selected_rows = st.session_state.selected_rows.union(yellow_indices)
        
        # Создаем пользовательскую таблицу с чекбоксами и кнопками PDF
        cols = st.columns([0.1, 1.5, 0.8, 0.8, 1.5, 0.8, 0.8, 0.8, 2, 0.3])
        cols[0].write("") # Пустой заголовок для чекбоксов
        cols[1].write("ФИО")
        cols[2].write("Вероятность")
        cols[3].write("Возраст")
        cols[4].write("Телефон")
        cols[5].write("Город")
        cols[6].write("Пол")
        cols[7].write("Зарплата")
        cols[8].write("Комментарий")
        cols[9].write("") # Пустой заголовок для кнопок PDF
        
        # Создаем строки для каждого резюме
        display_df = result_df.drop(columns=["raw_proba", "raw_text", "prediction_class"], errors="ignore")
        
        for idx, row in display_df.iterrows():
            # Определяем цвет фона строки в зависимости от вероятности
            try:
                prob = float(row["Вероятность класса 1"].replace(',', '.'))
            except:
                prob = 0
            
            if prob >= 0.81:
                bg_color = "#ccffcc"  # зеленый
            elif prob < 0.19:
                bg_color = "#ffcccc"  # красный
            else:
                bg_color = "#fff6cc"  # желтый
            
            # Создаем строку для каждого резюме
            row_cols = st.columns([0.1, 1.5, 0.8, 0.8, 1.5, 0.8, 0.8, 0.8, 2, 0.3])
            
            # Чекбокс для выбора
            is_selected = idx in st.session_state.selected_rows
            if row_cols[0].checkbox("", key=f"select_{idx}", value=is_selected):
                st.session_state.selected_rows.add(idx)
            else:
                if idx in st.session_state.selected_rows:
                    st.session_state.selected_rows.remove(idx)
            
            # Данные о кандидате
            name = row["Файл"].replace(".pdf", "")
            row_cols[1].markdown(f'<div style="background-color:{bg_color}; padding:5px;">{name}</div>', unsafe_allow_html=True)
            row_cols[2].markdown(f'<div style="background-color:{bg_color}; padding:5px;">{row["Вероятность класса 1"]}</div>', unsafe_allow_html=True)
            row_cols[3].markdown(f'<div style="background-color:{bg_color}; padding:5px;">{row["Возраст"]}</div>', unsafe_allow_html=True)
            row_cols[4].markdown(f'<div style="background-color:{bg_color}; padding:5px;">{row["Телефон"]}</div>', unsafe_allow_html=True)
            row_cols[5].markdown(f'<div style="background-color:{bg_color}; padding:5px;">{row["Город"]}</div>', unsafe_allow_html=True)
            row_cols[6].markdown(f'<div style="background-color:{bg_color}; padding:5px;">{row["Пол"]}</div>', unsafe_allow_html=True)
            row_cols[7].markdown(f'<div style="background-color:{bg_color}; padding:5px;">{row["Зарплата"]}</div>', unsafe_allow_html=True)
            
            # Комментарий с ограничением длины для улучшения отображения
            comment = row["Комментарий"]
            row_cols[8].markdown(f'<div style="background-color:{bg_color}; padding:5px; overflow-wrap: break-word;">{comment}</div>', unsafe_allow_html=True)
            
            # Кнопка PDF
            row_cols[9].markdown(f"""
            <style>
                div[data-testid="stButton"] > button {{
                    background-color: transparent;
                    color: #FF0000;
                    font-size: 12px;
                    padding: 1px 6px;
                    border: 1px solid #FF0000;
                    border-radius: 4px;
                    width: 80%;
                    height: auto;
                    margin: 0 auto;
                    display: block;
                }}
            </style>
            """, unsafe_allow_html=True)
            if row_cols[9].button("PDF", key=f"pdf_{idx}", help="Просмотр резюме"):
                # Найдем имя файла в processed_files
                file_name = row["Файл"] + ".pdf" if not row["Файл"].endswith(".pdf") else row["Файл"]
                if file_name in st.session_state.processed_files:
                    file_data = st.session_state.processed_files[file_name]["file"]
                    st.session_state.selected_pdf = {"file": file_data, "name": file_name}
                    st.rerun()  # Перезагрузить страницу для отображения PDF
        

        # Создаем буфер для Excel файла
        buffer = io.BytesIO()

        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Результаты анализа"

        # Добавляем заголовки
        headers = list(display_df.columns)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            
            # Настройка ширины столбца
            ws.column_dimensions[get_column_letter(col_num)].width = 15

        # Записываем данные
        for row_idx, row in enumerate(display_df.values, 2):
            # Получаем вероятность для определения цвета строки
            try:
                prob = float(row[list(display_df.columns).index("Вероятность класса 1")].replace(',', '.'))
            except:
                prob = 0
            
            # Определяем цвет фона в зависимости от вероятности
            if prob >= 0.81:
                fill_color = "CCFFCC"  # зеленый
            elif prob < 0.19:
                fill_color = "FFCCCC"  # красный
            else:
                fill_color = "FFF6CC"  # желтый
                
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            # Добавляем данные в строку
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.fill = fill
                
                # Для комментариев устанавливаем перенос текста
                if col_idx == list(display_df.columns).index("Комментарий") + 1:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    ws.row_dimensions[row_idx].height = 60  # Увеличиваем высоту строки

        # Сохраняем в буфер
        wb.save(buffer)
        buffer.seek(0)

        # Создаем кнопку скачивания Excel
        st.download_button(
            label="Скачать результаты (Excel)",
            data=buffer,
            file_name="predictions.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        
        st.markdown("""
        <style>
            div[data-testid="stButton"] > button {
                border-color: #9e9e9e;
                color: #505050;
            }
            div[data-testid="stButton"] > button:hover {
                border-color: #6e6e6e;
                color: #303030;
            }
        </style>
        """, unsafe_allow_html=True)


        # Кнопка для полной очистки всех данных резюме
        if st.button("Очистить все выбранные резюме", key="clear_all_button"):
            # JavaScript для перехода на URL с параметром очистки
            js = """
            <script>
                var url = window.parent.location.origin + window.parent.location.pathname;
                window.parent.location.href = url + "?clear=true&t=" + new Date().getTime();
            </script>
            """
            st.markdown(js, unsafe_allow_html=True)
            
            
            # Обязательно инициализируем заново с пустыми значениями
            st.session_state.processed_files = {}
            st.session_state.results = []
            st.session_state.has_processed_files = False
            st.session_state.selected_rows = set()


                
            # Сообщение об успешной очистке
            st.success("Все резюме успешно очищены!")
            
            # Принудительно перезагружаем страницу
            st.rerun()  # Используем просто rerun() вместо experimental_rerun()
        
        # Добавляем информацию о количестве выбранных резюме
        if st.session_state.selected_rows:
            selected_count = len(st.session_state.selected_rows)
            st.write(f"Выбрано: {selected_count} резюме")
            
            # Кнопка для отправки выбранных резюме в AmoCRM
            if st.button(f"Отправить выбранные резюме в AmoCRM ({selected_count})"):
                selected_results = [st.session_state.results[idx] for idx in st.session_state.selected_rows]
                
                # Создаем временный DataFrame только с выбранными резюме
                temp_df = pd.DataFrame(selected_results)
                temp_df["Файл"] = temp_df["Файл"].str.replace('.pdf', '', regex=False)
                temp_df["Вероятность класса 1"] = temp_df["Вероятность класса 1"].astype(float).map("{:.2f}".format)
                temp_df["Зарплата"] = temp_df["Зарплата"].apply(lambda x: f"{int(x):,}".replace(',', ' ') if str(x).isdigit() else x)
                
                display_temp_df = temp_df.drop(columns=["raw_proba", "raw_text", "prediction_class"], errors="ignore")
                
                # Сохраняем DataFrame в CSV
                temp_csv_path = "temp_selected_results.csv"
                display_temp_df.to_csv(temp_csv_path, index=False, encoding='utf-8')
                
                try:
                    # Вызываем функцию из amo_script.py
                    from amo_script import AmoCRMClient
                    
                    with st.spinner("Отправка выбранных данных в AmoCRM..."):
                        client = AmoCRMClient(temp_csv_path)
                        client.process_csv()
                    
                    # Удаляем временный CSV файл
                    if os.path.exists(temp_csv_path):
                        os.remove(temp_csv_path)
                    
                    st.success(f"Выбранные резюме ({selected_count}) успешно отправлены в AmoCRM!")
                except Exception as e:
                    st.error(f"Ошибка при отправке данных в AmoCRM: {e}")
        else:
            # Обычная кнопка отправки всех данных в AmoCRM
            if st.button("Отправить все данные в AmoCRM"):
                success = send_to_amocrm()
                if success:
                    st.success("Данные успешно отправлены в AmoCRM!")
        
        # Если выбран PDF для просмотра, отображаем его
        if hasattr(st.session_state, 'selected_pdf') and st.session_state.selected_pdf:
            st.divider()
            st.subheader(f"📄 Просмотр: {st.session_state.selected_pdf['name']}")
            display_pdf(st.session_state.selected_pdf['file'])
            
            # Информация о кандидате
            file_name = st.session_state.selected_pdf['name']
            raw_text = st.session_state.processed_files[file_name]["raw_text"]
            info = extract_resume_info(raw_text)
            info_df = pd.DataFrame({
                "Поле": ["Телефон", "Должность", "Город", "Возраст", "Пол", "Зарплата"],
                "Значение": [
                    info["phone"], 
                    info["position"], 
                    info["city"], 
                    info["age"], 
                    info["gender"],
                    info["salary"]
                ]
            })
            st.table(info_df)
            
            # Кнопка закрыть просмотр PDF
            if st.button("Закрыть просмотр PDF"):
                del st.session_state.selected_pdf
                st.rerun()
    else:
        if uploaded_files:
            st.info("Нажмите кнопку 'Обработать файлы' для анализа резюме.")
        else:
            st.info("Загрузите PDF-файлы резюме для анализа.")
    
    # --- Загрузка резюме с почты ---
    if st.session_state.user_role == "admin":
        st.divider()
        st.subheader("Загрузка резюме с почты")
        if st.button("Получить резюме с почты"):
            try:
                from pochtalion import download_pdfs
                downloaded_files = download_pdfs()
                if downloaded_files:
                    # Обработка новых файлов
                    results = []
                    for file_path in downloaded_files:
                        with open(file_path, "rb") as f:
                            file = io.BytesIO(f.read())
                            file.name = os.path.basename(file_path)
                            raw_text = extract_text_from_pdf(file)
                            info = extract_resume_info(raw_text)
                            processed_text = preprocess_resume(raw_text)
                            keyword_features = extract_features(processed_text, features)
                            resume_features = extract_resume_features(raw_text)
                            manual_df = pd.DataFrame([keyword_features | resume_features])
                            tfidf_features = tfidf.transform([processed_text]).toarray()
                            combined_features = np.hstack([manual_df.values, tfidf_features])
                            scaled_features = scaler.transform(combined_features)
                            proba = model.predict_proba(scaled_features)[0]
                            raw_proba = proba[1]
                            prediction = 1 if raw_proba >= THRESHOLD else 0
                            
                            comment, is_red_flag = get_detailed_comment(raw_text, prediction, raw_proba)
                            
                            results.append({
                                "Файл": file.name,
                                "Вероятность класса 1": raw_proba,
                                "Телефон": info["phone"],
                                "Желаемая должность": info["position"],
                                "Город": info["city"],
                                "Возраст": info["age"],
                                "Пол": info["gender"],
                                "Зарплата": info["salary"],
                                "Комментарий": comment,
                                "raw_proba": raw_proba,
                                "raw_text": raw_text
                            })
                            
                            # Сохраняем файл в processed_files для возможности просмотра PDF
                            st.session_state.processed_files[file.name] = {
                                "file": file,
                                "raw_text": raw_text
                            }
                            
                    st.session_state.results.extend(results)
                    st.session_state.has_processed_files = True
                    st.success(f"Загружено {len(downloaded_files)} новых резюме")
                    st.rerun()  # Перезагружаем страницу для отображения результатов
                else:
                    st.info("Новых резюме не найдено")
            except Exception as e:
                st.error(f"Ошибка при загрузке резюме с почты: {e}")
                st.info("Убедитесь, что файл pochtalion.py находится в той же директории и содержит функцию download_pdfs")
# --- Главная ---
# В функции main_app(), после отображения таблицы резюме и 
# перед блоком "Добавляем кнопки для действий с выбранными резюме":

# --- Главная ---
# --- Главная ---
def main():
    if st.session_state.authenticated:
        st.sidebar.write(f"Вы вошли как: **{st.session_state.user_name}** ({st.session_state.user_role})")
        
        # Добавляем меню в сайдбар
        menu_options = ["Классификация резюме"]
        
        # Дополнительный пункт меню только для администратора
        if st.session_state.user_role == "admin":
            menu_options.append("Панель администратора")
        
        selected_menu = st.sidebar.radio("Меню", menu_options)
        
        if st.sidebar.button("Выйти"):
            # Полный сброс всех ключей сессии
            if 'processed_files' in st.session_state:
                del st.session_state['processed_files']
            if 'results' in st.session_state:
                del st.session_state['results']
            if 'has_processed_files' in st.session_state:
                del st.session_state['has_processed_files']
            if 'selected_rows' in st.session_state:
                del st.session_state['selected_rows']
            if 'selected_pdf' in st.session_state:
                del st.session_state['selected_pdf']
            
            # Очистка авторизации
            st.session_state.authenticated = False
            st.session_state.user_role = None
            st.session_state.user_name = None
            
            # Инициализируем пустые структуры данных
            st.session_state.processed_files = {}
            st.session_state.results = []
            st.session_state.has_processed_files = False
            st.session_state.selected_rows = set()
            
            st.rerun()
        
        # Отображение соответствующей страницы в зависимости от выбора
        if selected_menu == "Классификация резюме":
            main_app()
        elif selected_menu == "Панель администратора" and st.session_state.user_role == "admin":
            admin_panel()
    else:
        login_page()

if __name__ == "__main__":
    main()